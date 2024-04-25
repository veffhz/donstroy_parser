import datetime
import math
import os

import requests
from openpyxl import Workbook
from openpyxl import load_workbook

from request import data_request

REPORT_XLSX = 'report.xlsx'

BASE_URL = 'https://donstroy.moscow/api/v1/flatssearch/choose_params_api_flats/'


def get_page(page_no=1):
    data_request['page'] = page_no
    response = requests.post(BASE_URL, json=data_request)
    return response.json()


def main(is_log=False):
    data = get_page()

    total_flats = data['total_flats']

    flats = [flat for flat in data['flats'] if not flat['isUtp']]

    pages_count = math.ceil(total_flats / (len(flats)))

    now = datetime.datetime.today().strftime("%Y-%m-%d_%H-%M-%S")

    if not os.path.isfile(REPORT_XLSX):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = now
    else:
        workbook = load_workbook(REPORT_XLSX)
        worksheet = workbook.create_sheet(now)

    row = 1
    column = 1

    for page in range(1, pages_count + 1):

        if page > 1:
            data = get_page(page)
            flats = [flat for flat in data['flats'] if not flat['isUtp']]

        for flat in flats:

            rooms = f'{flat["rooms"]}-комнатная квартира {flat["number"]}'
            project = f'{flat["project"]}'
            details = f'{flat["quarter"]}, Корпус {flat["building"]}, секция {flat["section"]}, этаж {flat["floor"]}'

            if is_log:
                template = rooms + '\n' + project + '\n' + details

                print(template)
                print()

            worksheet.cell(row, column, rooms)
            worksheet.cell(row, column + 1, project)
            worksheet.cell(row, column + 2, details)

            row += 1

    workbook.save(REPORT_XLSX)


if __name__ == "__main__":
    main()
